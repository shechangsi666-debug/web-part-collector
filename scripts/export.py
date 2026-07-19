# -*- coding: utf-8 -*-
"""Export collected products to Excel/CSV and organize product images."""

import csv
import os
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
except ImportError:
    openpyxl = None

HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
DATA_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

EXCEL_COLUMNS = [
    "主机类型",
    "主机品牌",
    "主机型号",
    "配件名称(标准)",
    "官网原始名称",
    "配件件号",
    "产品描述",
    "官网链接",
    "图片名称",
    "图片路径",
    "采集时间",
    "备注",
]

IMAGE_CATEGORIES = [
    "支重轮",
    "托链轮",
    "引导轮",
    "驱动齿",
    "链条",
    "履带板",
    "履带总成",
    "涨紧总成",
    "待分类",
]


def export_to_excel(products, output_path):
    if openpyxl is None:
        csv_path = output_path.replace(".xlsx", ".csv")
        export_to_csv(products, csv_path)
        return

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "产品数据"

    for col, header in enumerate(EXCEL_COLUMNS, 1):
        cell = sheet.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = DATA_ALIGN
        cell.border = THIN_BORDER

    for row_index, product in enumerate(products, 2):
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        values = [
            product.get("machine_type", ""),
            product.get("brand", ""),
            product.get("model", ""),
            product.get("std_name", ""),
            product.get("raw_name", ""),
            product.get("part_no", ""),
            product.get("description", ""),
            product.get("url", ""),
            product.get("image_name", ""),
            product.get("image_path", ""),
            product.get("timestamp", timestamp),
            product.get("notes", ""),
        ]
        for col, value in enumerate(values, 1):
            cell = sheet.cell(row=row_index, column=col, value=value)
            cell.alignment = DATA_ALIGN
            cell.border = THIN_BORDER

    workbook.save(output_path)
    print(f"  Excel: {output_path} ({len(products)} rows)")


def export_to_csv(products, output_path):
    with open(output_path, "w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.writer(handle)
        writer.writerow(EXCEL_COLUMNS)
        for product in products:
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            writer.writerow(
                [
                    product.get("machine_type", ""),
                    product.get("brand", ""),
                    product.get("model", ""),
                    product.get("std_name", ""),
                    product.get("raw_name", ""),
                    product.get("part_no", ""),
                    product.get("description", ""),
                    product.get("url", ""),
                    product.get("image_name", ""),
                    product.get("image_path", ""),
                    product.get("timestamp", timestamp),
                    product.get("notes", ""),
                ]
            )
    print(f"  CSV: {output_path} ({len(products)} rows)")


def organize_images(products, base_dir):
    import shutil

    organized = {category: [] for category in IMAGE_CATEGORIES}
    for product in products:
        category = product.get("std_name", "") or "待分类"
        if category not in organized:
            category = "待分类"

        source_path = product.get("image_path", "")
        if not source_path or not os.path.exists(source_path):
            continue

        destination_dir = os.path.join(base_dir, category)
        os.makedirs(destination_dir, exist_ok=True)
        filename = os.path.basename(source_path)
        destination_path = os.path.join(destination_dir, filename)
        try:
            shutil.copy2(source_path, destination_path)
            organized[category].append(destination_path)
            product["image_path"] = destination_path
        except Exception:
            pass
    return organized
